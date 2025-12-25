"""
Equipment Control Relationship Template Generator
==================================================

설비 제어 참조 관계 리스트 작성을 위한 엑셀 템플릿
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


def create_template():
    """설비 제어 관계 템플릿 생성"""

    wb = Workbook()

    # 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_fill = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    setpoint_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    pv_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    quality_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ================================================================
    # Sheet 1: 설비 마스터 (Equipment Master)
    # ================================================================
    ws1 = wb.active
    ws1.title = "1.설비마스터"

    headers1 = [
        ("A", "No", 5),
        ("B", "설비 ID", 15),
        ("C", "설비명", 20),
        ("D", "설비 유형", 15),
        ("E", "공정 단계", 15),
        ("F", "제조사", 15),
        ("G", "모델명", 20),
        ("H", "챔버 수", 10),
        ("I", "데이터 소스\n(FDC 테이블)", 20),
        ("J", "비고", 25),
    ]

    for col, header, width in headers1:
        cell = ws1[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws1.column_dimensions[col].width = width

    examples1 = [
        (1, "ETCH-001", "Etcher #1", "DRY_ETCH", "ETCH", "LAM", "Kiyo CX", 2, "FDC_ETCH_001", "주력 설비"),
        (2, "ETCH-002", "Etcher #2", "DRY_ETCH", "ETCH", "LAM", "Kiyo CX", 2, "FDC_ETCH_002", ""),
        (3, "CVD-001", "CVD #1", "CVD", "DEP", "AMAT", "Producer GT", 4, "FDC_CVD_001", ""),
        (4, "LITHO-001", "Scanner #1", "LITHO", "PHOTO", "ASML", "NXT:1980", 1, "FDC_LITHO_001", "EUV"),
    ]

    for i, example in enumerate(examples1, start=2):
        for j, value in enumerate(example):
            cell = ws1.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = center_align if j < 8 else left_align

    # 빈 행
    for i in range(6, 50):
        ws1.cell(row=i, column=1, value=i-1)
        for j in range(1, 11):
            ws1.cell(row=i, column=j).border = thin_border

    # ================================================================
    # Sheet 2: Setpoint (제어 설정값)
    # ================================================================
    ws2 = wb.create_sheet("2.Setpoint(설정값)")

    headers2 = [
        ("A", "No", 5),
        ("B", "설비 ID", 15),
        ("C", "파라미터 ID", 20),
        ("D", "파라미터명\n(한글)", 20),
        ("E", "단위", 10),
        ("F", "제어 대상\n(Actuator)", 20),
        ("G", "기본값", 12),
        ("H", "허용 범위\n(Min)", 12),
        ("I", "허용 범위\n(Max)", 12),
        ("J", "제어 방식", 15),
        ("K", "영향받는 PV", 25),
        ("L", "비고", 20),
    ]

    for col, header, width in headers2:
        cell = ws2[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws2.column_dimensions[col].width = width

    examples2 = [
        (1, "ETCH-001", "RF_POWER_SP", "RF 파워 설정", "W", "RF Generator", 500, 300, 700, "Direct", "RF_POWER_PV, PLASMA_DENSITY", ""),
        (2, "ETCH-001", "PRESSURE_SP", "압력 설정", "mTorr", "Throttle Valve", 10, 5, 20, "PID", "PRESSURE_PV, GAS_RESIDENCE", ""),
        (3, "ETCH-001", "GAS1_FLOW_SP", "가스1 유량 설정", "sccm", "MFC-1", 100, 50, 200, "PID", "GAS1_FLOW_PV, PRESSURE_PV", "CF4"),
        (4, "ETCH-001", "GAS2_FLOW_SP", "가스2 유량 설정", "sccm", "MFC-2", 50, 20, 100, "PID", "GAS2_FLOW_PV, PRESSURE_PV", "O2"),
        (5, "ETCH-001", "ESC_TEMP_SP", "척 온도 설정", "°C", "Chiller", 60, 40, 80, "PID", "ESC_TEMP_PV, WAFER_TEMP", ""),
        (6, "ETCH-001", "BIAS_POWER_SP", "바이어스 파워", "W", "Bias RF Gen", 200, 100, 400, "Direct", "BIAS_POWER_PV, ION_ENERGY", ""),
    ]

    for i, example in enumerate(examples2, start=2):
        for j, value in enumerate(example):
            cell = ws2.cell(row=i, column=j+1, value=value)
            cell.fill = setpoint_fill
            cell.border = thin_border
            cell.alignment = center_align if j < 10 else left_align

    for i in range(8, 100):
        for j in range(1, 13):
            ws2.cell(row=i, column=j).border = thin_border

    # ================================================================
    # Sheet 3: Process Variable (측정값)
    # ================================================================
    ws3 = wb.create_sheet("3.ProcessVariable(측정값)")

    headers3 = [
        ("A", "No", 5),
        ("B", "설비 ID", 15),
        ("C", "파라미터 ID", 20),
        ("D", "파라미터명\n(한글)", 20),
        ("E", "단위", 10),
        ("F", "센서/측정 방식", 20),
        ("G", "샘플링\n주기(초)", 10),
        ("H", "정상 범위\n(Min)", 12),
        ("I", "정상 범위\n(Max)", 12),
        ("J", "제어되는\nSetpoint", 20),
        ("K", "영향주는\n다른 PV", 25),
        ("L", "영향받는\n다른 PV", 25),
        ("M", "품질 영향", 20),
    ]

    for col, header, width in headers3:
        cell = ws3[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws3.column_dimensions[col].width = width

    examples3 = [
        (1, "ETCH-001", "RF_POWER_PV", "RF 파워 실측", "W", "Power Meter", 1, 480, 520, "RF_POWER_SP", "", "PLASMA_DENSITY, ETCH_RATE", "Etch Rate"),
        (2, "ETCH-001", "RF_REFLECT_PV", "RF 반사파", "W", "Power Meter", 1, 0, 50, "", "RF_POWER_PV", "", "Uniformity"),
        (3, "ETCH-001", "PRESSURE_PV", "챔버 압력 실측", "mTorr", "Baratron", 1, 9, 11, "PRESSURE_SP", "GAS_FLOW_PV", "PLASMA_DENSITY", "Profile"),
        (4, "ETCH-001", "GAS1_FLOW_PV", "가스1 유량 실측", "sccm", "MFC Feedback", 1, 95, 105, "GAS1_FLOW_SP", "", "PRESSURE_PV", ""),
        (5, "ETCH-001", "PLASMA_DENSITY", "플라즈마 밀도", "a.u.", "OES", 1, 0.8, 1.2, "", "RF_POWER_PV, PRESSURE_PV", "ETCH_RATE", "Etch Rate"),
        (6, "ETCH-001", "ETCH_RATE", "식각률", "A/min", "Endpoint Det.", 1, 900, 1100, "", "PLASMA_DENSITY, BIAS_POWER", "", "CD, Depth"),
        (7, "ETCH-001", "ESC_TEMP_PV", "척 온도 실측", "°C", "Thermocouple", 1, 58, 62, "ESC_TEMP_SP", "", "WAFER_TEMP", "Uniformity"),
        (8, "ETCH-001", "WAFER_TEMP", "웨이퍼 온도", "°C", "Pyrometer", 1, 50, 70, "", "ESC_TEMP_PV", "ETCH_RATE", "Profile"),
    ]

    for i, example in enumerate(examples3, start=2):
        for j, value in enumerate(example):
            cell = ws3.cell(row=i, column=j+1, value=value)
            cell.fill = pv_fill
            cell.border = thin_border
            cell.alignment = center_align if j < 9 else left_align

    for i in range(10, 150):
        for j in range(1, 14):
            ws3.cell(row=i, column=j).border = thin_border

    # ================================================================
    # Sheet 4: 제어 루프 (Control Loop)
    # ================================================================
    ws4 = wb.create_sheet("4.제어루프")

    headers4 = [
        ("A", "No", 5),
        ("B", "설비 ID", 15),
        ("C", "루프 ID", 15),
        ("D", "루프명", 25),
        ("E", "Setpoint", 20),
        ("F", "Process Variable", 20),
        ("G", "Actuator", 20),
        ("H", "제어 알고리즘", 15),
        ("I", "응답 시간(초)", 12),
        ("J", "안정화 시간(초)", 12),
        ("K", "Interlock 조건", 30),
    ]

    for col, header, width in headers4:
        cell = ws4[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws4.column_dimensions[col].width = width

    examples4 = [
        (1, "ETCH-001", "LOOP-PRESS", "압력 제어 루프", "PRESSURE_SP", "PRESSURE_PV", "Throttle Valve", "PID", 0.5, 3, "PRESSURE > 50mTorr → Abort"),
        (2, "ETCH-001", "LOOP-GAS1", "가스1 유량 제어", "GAS1_FLOW_SP", "GAS1_FLOW_PV", "MFC-1", "PID", 0.2, 1, "FLOW > 300sccm → Alarm"),
        (3, "ETCH-001", "LOOP-TEMP", "척 온도 제어", "ESC_TEMP_SP", "ESC_TEMP_PV", "Chiller", "PID", 5, 60, "TEMP > 100°C → Shutdown"),
        (4, "ETCH-001", "LOOP-RF", "RF 파워 제어", "RF_POWER_SP", "RF_POWER_PV", "RF Generator", "Direct", 0.1, 0.5, "REFLECT > 100W → Retune"),
    ]

    for i, example in enumerate(examples4, start=2):
        for j, value in enumerate(example):
            cell = ws4.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = center_align if j < 8 else left_align

    for i in range(6, 50):
        for j in range(1, 12):
            ws4.cell(row=i, column=j).border = thin_border

    # ================================================================
    # Sheet 5: PV 간 인과관계 (PV Causality)
    # ================================================================
    ws5 = wb.create_sheet("5.PV간_인과관계")

    headers5 = [
        ("A", "No", 5),
        ("B", "설비 ID", 15),
        ("C", "원인 PV", 20),
        ("D", "결과 PV", 20),
        ("E", "관계 유형", 15),
        ("F", "방향성", 12),
        ("G", "시간 지연\n(초)", 12),
        ("H", "영향 강도\n(0~1)", 12),
        ("I", "물리적 메커니즘", 40),
        ("J", "검증 방법", 20),
    ]

    for col, header, width in headers5:
        cell = ws5[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws5.column_dimensions[col].width = width

    examples5 = [
        (1, "ETCH-001", "RF_POWER_PV", "PLASMA_DENSITY", "CAUSES", "positive", 0.1, 0.9, "RF 에너지가 가스를 이온화하여 플라즈마 생성", "OES 측정"),
        (2, "ETCH-001", "PLASMA_DENSITY", "ETCH_RATE", "CAUSES", "positive", 0.5, 0.85, "플라즈마 이온이 표면과 반응하여 식각", "EPD 측정"),
        (3, "ETCH-001", "PRESSURE_PV", "PLASMA_DENSITY", "CAUSES", "negative", 0.2, 0.7, "압력 상승 시 평균자유행로 감소로 밀도 저하", "OES 측정"),
        (4, "ETCH-001", "GAS1_FLOW_PV", "PRESSURE_PV", "CAUSES", "positive", 2, 0.6, "가스 유입량 증가 시 배기 한계로 압력 상승", "Baratron"),
        (5, "ETCH-001", "ESC_TEMP_PV", "WAFER_TEMP", "CAUSES", "positive", 10, 0.8, "척에서 웨이퍼로 열전달", "Pyrometer"),
        (6, "ETCH-001", "WAFER_TEMP", "ETCH_RATE", "CAUSES", "positive", 1, 0.4, "온도 상승 시 반응 속도 증가 (Arrhenius)", "EPD"),
        (7, "ETCH-001", "BIAS_POWER_PV", "ION_ENERGY", "CAUSES", "positive", 0.1, 0.95, "바이어스가 이온 가속", "계산"),
        (8, "ETCH-001", "ION_ENERGY", "ETCH_RATE", "CAUSES", "positive", 0.2, 0.6, "이온 에너지가 물리적 스퍼터링 증가", "EPD"),
    ]

    for i, example in enumerate(examples5, start=2):
        for j, value in enumerate(example):
            cell = ws5.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = center_align if j < 8 else left_align

    for i in range(10, 100):
        for j in range(1, 11):
            ws5.cell(row=i, column=j).border = thin_border

    # ================================================================
    # Sheet 6: 품질 영향 관계 (Quality Impact)
    # ================================================================
    ws6 = wb.create_sheet("6.품질영향관계")

    headers6 = [
        ("A", "No", 5),
        ("B", "설비 ID", 15),
        ("C", "PV 파라미터", 20),
        ("D", "품질 지표", 20),
        ("E", "영향 방향", 12),
        ("F", "영향 강도\n(0~1)", 12),
        ("G", "허용 편차", 15),
        ("H", "편차 초과 시\n품질 영향", 30),
        ("I", "측정 단계", 15),
        ("J", "비고", 20),
    ]

    for col, header, width in headers6:
        cell = ws6[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws6.column_dimensions[col].width = width

    examples6 = [
        (1, "ETCH-001", "ETCH_RATE", "CD (Critical Dimension)", "positive", 0.9, "±5%", "CD 증가로 회로 단선 위험", "Metrology", "핵심 품질"),
        (2, "ETCH-001", "ETCH_RATE", "Depth", "positive", 0.95, "±3%", "Depth 부족 시 접촉 불량", "Metrology", ""),
        (3, "ETCH-001", "RF_REFLECT_PV", "Uniformity", "negative", 0.7, "<50W", "반사파 높으면 균일도 저하", "Metrology", ""),
        (4, "ETCH-001", "WAFER_TEMP", "Profile Angle", "positive", 0.6, "±2°C", "온도 불균일 시 기울기 발생", "SEM", ""),
        (5, "ETCH-001", "PRESSURE_PV", "Selectivity", "complex", 0.5, "±0.5mTorr", "압력 변화 시 선택비 변화", "Metrology", "레시피 의존"),
        (6, "ETCH-001", "PLASMA_DENSITY", "Etch Damage", "positive", 0.4, "±10%", "밀도 과다 시 기판 손상", "Electrical", ""),
    ]

    for i, example in enumerate(examples6, start=2):
        for j, value in enumerate(example):
            cell = ws6.cell(row=i, column=j+1, value=value)
            cell.fill = quality_fill
            cell.border = thin_border
            cell.alignment = center_align if j < 7 else left_align

    for i in range(8, 50):
        for j in range(1, 11):
            ws6.cell(row=i, column=j).border = thin_border

    # ================================================================
    # Sheet 7: 알람-원인 매핑 (Alarm Cause Mapping)
    # ================================================================
    ws7 = wb.create_sheet("7.알람원인매핑")

    headers7 = [
        ("A", "No", 5),
        ("B", "설비 ID", 15),
        ("C", "알람 코드", 20),
        ("D", "알람 설명", 30),
        ("E", "심각도", 10),
        ("F", "1순위 원인 PV", 20),
        ("G", "1순위 조건", 25),
        ("H", "2순위 원인 PV", 20),
        ("I", "2순위 조건", 25),
        ("J", "3순위 원인 PV", 20),
        ("K", "즉시 조치", 30),
    ]

    for col, header, width in headers7:
        cell = ws7[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws7.column_dimensions[col].width = width

    examples7 = [
        (1, "ETCH-001", "ALM-RF-001", "RF 반사파 과다", "MAJOR", "RF_REFLECT_PV", "> 100W", "RF_POWER_PV", "변동 > 10%", "PRESSURE_PV", "Auto Matching 실행"),
        (2, "ETCH-001", "ALM-PRESS-001", "압력 이탈", "MAJOR", "THROTTLE_POS", "Full Open/Close", "GAS_FLOW_PV", "설정 대비 ±20%", "PUMP_CURRENT", "가스/펌프 점검"),
        (3, "ETCH-001", "ALM-TEMP-001", "온도 이탈", "MINOR", "CHILLER_TEMP", "설정 대비 ±5°C", "ESC_TEMP_PV", "변동 > 2°C/min", "", "Chiller 점검"),
        (4, "ETCH-001", "ALM-EPD-001", "EPD 미검출", "CRITICAL", "ETCH_RATE", "< 100 A/min", "PLASMA_DENSITY", "< 0.3", "RF_POWER_PV", "공정 중단, 레시피 확인"),
        (5, "ETCH-001", "ALM-LEAK-001", "진공 누설", "CRITICAL", "PRESSURE_PV", "Baseline 상승", "PUMP_CURRENT", "정상 대비 +20%", "", "챔버 Leak 점검"),
    ]

    for i, example in enumerate(examples7, start=2):
        for j, value in enumerate(example):
            cell = ws7.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = center_align if j < 5 else left_align

    for i in range(7, 50):
        for j in range(1, 12):
            ws7.cell(row=i, column=j).border = thin_border

    # ================================================================
    # Sheet 8: 데이터 소스 매핑 (Data Source Mapping)
    # ================================================================
    ws8 = wb.create_sheet("8.데이터소스매핑")

    headers8 = [
        ("A", "No", 5),
        ("B", "파라미터 ID", 20),
        ("C", "소스 시스템", 15),
        ("D", "테이블/뷰 명", 25),
        ("E", "컬럼명", 20),
        ("F", "데이터 타입", 12),
        ("G", "수집 주기", 12),
        ("H", "DataWarehouse\n테이블", 25),
        ("I", "변환 로직", 30),
        ("J", "비고", 20),
    ]

    for col, header, width in headers8:
        cell = ws8[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws8.column_dimensions[col].width = width

    examples8 = [
        (1, "RF_POWER_PV", "FDC", "FDC_ETCH_001_PV", "RF_FWD_POWER", "FLOAT", "1초", "DW_FDC_ETCH", "단위 변환 없음", ""),
        (2, "PRESSURE_PV", "FDC", "FDC_ETCH_001_PV", "CHAMBER_PRESS", "FLOAT", "1초", "DW_FDC_ETCH", "Pa → mTorr 변환", "x 7.5"),
        (3, "ETCH_RATE", "FDC", "FDC_ETCH_001_EPD", "ETCH_RATE_CALC", "FLOAT", "Step 단위", "DW_FDC_ETCH", "", "계산값"),
        (4, "CD", "SPC", "SPC_METRO_CD", "CD_VALUE", "FLOAT", "Wafer 단위", "DW_SPC_QUALITY", "", "측정값"),
        (5, "ALARM_CODE", "EDA", "EDA_ALARM_HIST", "ALARM_ID", "VARCHAR", "이벤트", "DW_ALARM", "", ""),
    ]

    for i, example in enumerate(examples8, start=2):
        for j, value in enumerate(example):
            cell = ws8.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = center_align if j < 7 else left_align

    for i in range(7, 100):
        for j in range(1, 11):
            ws8.cell(row=i, column=j).border = thin_border

    # ================================================================
    # Sheet 9: 가이드
    # ================================================================
    ws9 = wb.create_sheet("가이드")

    guide = [
        ("설비 제어 참조 관계 리스트 작성 가이드", Font(bold=True, size=14), header_fill),
        ("", None, None),
        ("이 템플릿은 설비의 제어 구조와 데이터 간 관계를 체계적으로 정리하기 위한 것입니다.", None, None),
        ("작성된 내용은 AI Agent의 추론 기반이 되며, 이상 감지 및 근본 원인 분석에 활용됩니다.", None, None),
        ("", None, None),
        ("■ 시트 설명", Font(bold=True, size=12), subheader_fill),
        ("", None, None),
        ("1. 설비마스터: 분석 대상 설비 목록", None, None),
        ("   - 설비별 기본 정보와 데이터 소스 테이블 매핑", None, None),
        ("", None, None),
        ("2. Setpoint(설정값): 레시피에서 설정하는 제어 목표값", None, None),
        ("   - 제어하는 Actuator와 영향받는 PV 명시", None, None),
        ("   - 예: RF_POWER_SP → RF Generator → RF_POWER_PV", None, None),
        ("", None, None),
        ("3. ProcessVariable(측정값): 센서로 측정되는 실제 값", None, None),
        ("   - 제어되는 Setpoint, 영향 주고받는 다른 PV 명시", None, None),
        ("   - 품질에 미치는 영향 기록", None, None),
        ("", None, None),
        ("4. 제어루프: Setpoint-PV-Actuator 간 피드백 제어 구조", None, None),
        ("   - 응답 시간, 안정화 시간, Interlock 조건 기록", None, None),
        ("", None, None),
        ("5. PV간_인과관계: PV끼리의 물리적 인과관계 (핵심!)", None, None),
        ("   - 예: RF_POWER → PLASMA_DENSITY → ETCH_RATE", None, None),
        ("   - 시간 지연, 영향 강도, 물리적 메커니즘 기록", None, None),
        ("", None, None),
        ("6. 품질영향관계: PV와 최종 품질 지표 간 관계", None, None),
        ("   - 허용 편차 초과 시 품질에 미치는 영향 기록", None, None),
        ("", None, None),
        ("7. 알람원인매핑: 알람별 점검해야 할 원인 PV 우선순위", None, None),
        ("   - 즉시 조치 사항 기록", None, None),
        ("", None, None),
        ("8. 데이터소스매핑: 파라미터별 실제 데이터 소스 위치", None, None),
        ("   - DataWarehouse 테이블 매핑, 변환 로직 기록", None, None),
        ("", None, None),
        ("■ 작성 팁", Font(bold=True, size=12), subheader_fill),
        ("", None, None),
        ("• 초록색 행은 예시입니다. 실제 설비에 맞게 수정하세요.", None, None),
        ("• 모든 파라미터를 채울 필요 없습니다. 핵심 파라미터부터 시작하세요.", None, None),
        ("• 확실한 관계부터 작성하고, 불확실한 것은 나중에 데이터로 검증하세요.", None, None),
        ("• 한 설비를 완벽하게 작성한 후 다른 설비로 확장하세요.", None, None),
        ("• 현장 엔지니어와 함께 작성하면 가장 정확합니다.", None, None),
        ("", None, None),
        ("■ 관계 유형 정의", Font(bold=True, size=12), subheader_fill),
        ("", None, None),
        ("• CAUSES: 직접적 원인-결과 (A 변화 → B 변화 필연)", None, None),
        ("• INFLUENCES: 간접적 영향 (A 변화 → B 변화 가능성 증가)", None, None),
        ("• CORRELATES: 상관관계만 있음 (인과 방향 불명확)", None, None),
        ("• INHIBITS: 억제 관계 (A 증가 → B 감소)", None, None),
    ]

    ws9.column_dimensions['A'].width = 100

    for i, (text, font, fill) in enumerate(guide, start=1):
        cell = ws9.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill

    # ================================================================
    # 저장
    # ================================================================
    output_dir = os.path.join(os.path.dirname(__file__), '..', 'templates')
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, 'equipment_control_relationship_template.xlsx')
    wb.save(output_path)

    print(f"설비 제어 관계 템플릿 생성 완료: {output_path}")
    return output_path


if __name__ == '__main__':
    create_template()
