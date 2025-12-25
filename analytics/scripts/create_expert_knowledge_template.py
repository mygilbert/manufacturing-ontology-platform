"""
Expert Knowledge Excel Template Generator
==========================================

전문가 도메인 지식 입력을 위한 엑셀 템플릿 생성
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment


def create_template():
    """전문가 지식 입력 엑셀 템플릿 생성"""

    wb = Workbook()

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    warning_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ================================================================
    # Sheet 1: 인과관계 (Causal Relationships)
    # ================================================================
    ws1 = wb.active
    ws1.title = "1.인과관계"

    # 헤더
    headers1 = [
        ("A", "No", 5),
        ("B", "원인 파라미터\n(Source)", 20),
        ("C", "결과 파라미터\n(Target)", 20),
        ("D", "관계 유형", 15),
        ("E", "방향성", 12),
        ("F", "시간지연(초)\n최소", 12),
        ("G", "시간지연(초)\n최대", 12),
        ("H", "신뢰도\n(0~1)", 10),
        ("I", "물리적 설명", 50),
        ("J", "비고", 20),
    ]

    for col, header, width in headers1:
        cell = ws1[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws1.column_dimensions[col].width = width

    # 예시 데이터
    examples1 = [
        (1, "RF_Power", "Etch_Rate", "CAUSES", "positive", 1, 5, 1.0,
         "RF 파워가 플라즈마 밀도를 높여 식각 속도 증가", "핵심 관계"),
        (2, "Chamber_Pressure", "Plasma_Density", "CAUSES", "negative", 0, 2, 1.0,
         "압력 증가 시 평균자유행로 감소로 플라즈마 밀도 하락", ""),
        (3, "Gas_Flow_Rate", "Chamber_Pressure", "CAUSES", "positive", 2, 10, 0.9,
         "가스 유량 증가 시 배기 용량 초과하면 압력 상승", ""),
        (4, "ESC_Temperature", "Wafer_Temperature", "CAUSES", "positive", 5, 30, 0.95,
         "정전척 온도가 웨이퍼로 열전달", ""),
        (5, "Throttle_Position", "Chamber_Pressure", "CAUSES", "negative", 1, 5, 1.0,
         "스로틀 열림 → 배기 증가 → 압력 감소", ""),
    ]

    for i, example in enumerate(examples1, start=2):
        for j, value in enumerate(example):
            cell = ws1.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = left_align if j >= 8 else center_align

    # 빈 행 추가 (입력용)
    for i in range(7, 50):
        ws1.cell(row=i, column=1, value=i-1)
        for j in range(1, 11):
            cell = ws1.cell(row=i, column=j)
            cell.border = thin_border

    # 드롭다운 추가
    relation_types = DataValidation(type="list", formula1='"CAUSES,INFLUENCES,CORRELATES,INHIBITS"')
    direction_types = DataValidation(type="list", formula1='"positive,negative,complex"')
    ws1.add_data_validation(relation_types)
    ws1.add_data_validation(direction_types)
    relation_types.add(f"D2:D100")
    direction_types.add(f"E2:E100")

    # ================================================================
    # Sheet 2: 알람 원인 (Alarm Causes)
    # ================================================================
    ws2 = wb.create_sheet("2.알람원인")

    headers2 = [
        ("A", "No", 5),
        ("B", "알람 코드", 20),
        ("C", "알람 설명", 30),
        ("D", "원인 파라미터", 20),
        ("E", "이상 조건", 25),
        ("F", "발생 확률\n(0~1)", 12),
        ("G", "조치 방법", 40),
        ("H", "우선순위", 10),
    ]

    for col, header, width in headers2:
        cell = ws2[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws2.column_dimensions[col].width = width

    examples2 = [
        (1, "ETCH_RATE_OOS", "식각률 Spec Out", "RF_Power", "deviation > 5%", 0.7, "RF Power 점검 및 재설정", 1),
        (2, "ETCH_RATE_OOS", "식각률 Spec Out", "Chamber_Pressure", "out_of_spec", 0.5, "Throttle valve 점검", 2),
        (3, "ETCH_RATE_OOS", "식각률 Spec Out", "Gas_Flow_Rate", "unstable", 0.3, "MFC 점검", 3),
        (4, "PARTICLE_ALARM", "파티클 검출", "Chamber_Pressure", "sudden_change", 0.6, "챔버 클리닝", 1),
        (5, "PARTICLE_ALARM", "파티클 검출", "ESC_Temperature", "gradient > 5C/min", 0.4, "온도 안정화 대기", 2),
        (6, "UNIFORMITY_FAIL", "균일도 불량", "Gas_Flow_Ratio", "imbalance > 3%", 0.8, "가스 비율 조정", 1),
    ]

    for i, example in enumerate(examples2, start=2):
        for j, value in enumerate(example):
            cell = ws2.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = left_align if j >= 4 else center_align

    for i in range(8, 100):
        for j in range(1, 9):
            cell = ws2.cell(row=i, column=j)
            cell.border = thin_border

    # ================================================================
    # Sheet 3: 선행 지표 (Leading Indicators)
    # ================================================================
    ws3 = wb.create_sheet("3.선행지표")

    headers3 = [
        ("A", "No", 5),
        ("B", "예측 대상 이벤트", 25),
        ("C", "선행 지표 파라미터", 20),
        ("D", "패턴 유형", 20),
        ("E", "선행 시간(초)", 15),
        ("F", "임계값/조건", 25),
        ("G", "신뢰도\n(0~1)", 10),
        ("H", "설명", 40),
    ]

    for col, header, width in headers3:
        cell = ws3[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws3.column_dimensions[col].width = width

    examples3 = [
        (1, "EQUIPMENT_DOWN", "Vibration", "increasing_trend", 300, "기울기 > 0.1/min", 0.8, "진동 증가 추세 시 설비 고장 임박"),
        (2, "EQUIPMENT_DOWN", "Motor_Current", "spike", 60, "정상 대비 150%", 0.7, "전류 스파이크 발생 시 모터 이상"),
        (3, "QUALITY_FAIL", "Etch_Rate", "drift", 600, "3-sigma 이탈", 0.85, "식각률 드리프트 시 품질 이상"),
        (4, "QUALITY_FAIL", "Uniformity", "degradation", 1800, "Cpk < 1.0", 0.9, "균일도 저하 시 불량 발생"),
        (5, "PM_REQUIRED", "RF_Reflect", "gradual_increase", 86400, "초기 대비 20% 증가", 0.75, "반사파 증가 시 매칭 필요"),
    ]

    for i, example in enumerate(examples3, start=2):
        for j, value in enumerate(example):
            cell = ws3.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = left_align if j >= 5 else center_align

    pattern_types = DataValidation(type="list",
        formula1='"increasing_trend,decreasing_trend,spike,drift,oscillation,sudden_change,gradual_increase,gradual_decrease,degradation"')
    ws3.add_data_validation(pattern_types)
    pattern_types.add(f"D2:D100")

    for i in range(7, 50):
        for j in range(1, 9):
            cell = ws3.cell(row=i, column=j)
            cell.border = thin_border

    # ================================================================
    # Sheet 4: 파라미터 그룹 (Parameter Groups)
    # ================================================================
    ws4 = wb.create_sheet("4.파라미터그룹")

    headers4 = [
        ("A", "No", 5),
        ("B", "그룹명", 20),
        ("C", "설명", 30),
        ("D", "파라미터1", 18),
        ("E", "파라미터2", 18),
        ("F", "파라미터3", 18),
        ("G", "파라미터4", 18),
        ("H", "파라미터5", 18),
        ("I", "분석 시 주의사항", 35),
    ]

    for col, header, width in headers4:
        cell = ws4[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws4.column_dimensions[col].width = width

    examples4 = [
        (1, "플라즈마 상태", "플라즈마 특성 관련 파라미터", "RF_Power", "RF_Reflect", "Plasma_Density", "OES_Intensity", "", "RF Power와 Reflect는 함께 분석"),
        (2, "온도 제어", "온도 관련 파라미터", "ESC_Temp", "Wall_Temp", "Gas_Inlet_Temp", "Wafer_Temp", "", "온도 구배 확인 필요"),
        (3, "가스 시스템", "가스 공급 관련", "Gas_Flow_1", "Gas_Flow_2", "Gas_Flow_3", "Pressure", "Throttle", "가스비 변화 모니터링"),
        (4, "진공 시스템", "진공 상태 관련", "Pressure", "Pump_Current", "Throttle_Pos", "Leak_Rate", "", "압력 이상 시 전체 점검"),
    ]

    for i, example in enumerate(examples4, start=2):
        for j, value in enumerate(example):
            cell = ws4.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = left_align if j >= 2 else center_align

    for i in range(6, 30):
        for j in range(1, 10):
            cell = ws4.cell(row=i, column=j)
            cell.border = thin_border

    # ================================================================
    # Sheet 5: 불가능한 관계 (Impossible Relationships)
    # ================================================================
    ws5 = wb.create_sheet("5.불가능한관계")

    headers5 = [
        ("A", "No", 5),
        ("B", "파라미터1", 20),
        ("C", "파라미터2", 20),
        ("D", "불가능한 이유", 50),
        ("E", "비고", 30),
    ]

    for col, header, width in headers5:
        cell = ws5[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws5.column_dimensions[col].width = width

    examples5 = [
        (1, "Wafer_ID", "Etch_Rate", "Wafer ID는 식별자일 뿐 물리량이 아님", "상관관계 있어도 인과관계 아님"),
        (2, "Recipe_Name", "Chamber_Pressure", "Recipe는 설정값 묶음, 직접적 원인 아님", "Recipe 내 개별 파라미터로 분석"),
        (3, "Lot_ID", "Quality", "Lot ID는 그룹핑 정보", "Lot 내 조건으로 분석해야 함"),
        (4, "Timestamp", "Any_Parameter", "시간 자체는 원인이 아님", "시간에 따른 변화 패턴으로 분석"),
        (5, "Operator_ID", "Defect_Rate", "작업자 ID 자체는 원인 아님 (편향 주의)", "작업 패턴으로 분석"),
    ]

    for i, example in enumerate(examples5, start=2):
        for j, value in enumerate(example):
            cell = ws5.cell(row=i, column=j+1, value=value)
            cell.fill = warning_fill
            cell.border = thin_border
            cell.alignment = left_align if j >= 3 else center_align

    for i in range(7, 30):
        for j in range(1, 6):
            cell = ws5.cell(row=i, column=j)
            cell.border = thin_border

    # ================================================================
    # Sheet 6: 파라미터 정의 (Parameter Definitions)
    # ================================================================
    ws6 = wb.create_sheet("6.파라미터정의")

    headers6 = [
        ("A", "No", 5),
        ("B", "파라미터 ID", 20),
        ("C", "파라미터 한글명", 20),
        ("D", "단위", 10),
        ("E", "정상 범위\n(Min)", 12),
        ("F", "정상 범위\n(Max)", 12),
        ("G", "Spec\n(LSL)", 12),
        ("H", "Spec\n(USL)", 12),
        ("I", "데이터 소스", 15),
        ("J", "설명", 40),
    ]

    for col, header, width in headers6:
        cell = ws6[f"{col}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws6.column_dimensions[col].width = width

    examples6 = [
        (1, "RF_Power", "RF 파워", "W", 450, 550, 400, 600, "FDC", "플라즈마 생성용 RF 전력"),
        (2, "Chamber_Pressure", "챔버 압력", "mTorr", 8, 12, 5, 15, "FDC", "공정 챔버 내부 압력"),
        (3, "Etch_Rate", "식각률", "A/min", 900, 1100, 850, 1150, "FDC", "단위 시간당 식각 두께"),
        (4, "ESC_Temperature", "정전척 온도", "°C", 58, 62, 55, 65, "FDC", "웨이퍼 고정용 척 온도"),
        (5, "Gas_Flow_1", "가스1 유량", "sccm", 95, 105, 90, 110, "FDC", "공정 가스 1 유량"),
    ]

    for i, example in enumerate(examples6, start=2):
        for j, value in enumerate(example):
            cell = ws6.cell(row=i, column=j+1, value=value)
            cell.fill = example_fill
            cell.border = thin_border
            cell.alignment = left_align if j >= 9 else center_align

    for i in range(7, 100):
        for j in range(1, 11):
            cell = ws6.cell(row=i, column=j)
            cell.border = thin_border

    # ================================================================
    # Sheet 7: 사용 가이드
    # ================================================================
    ws7 = wb.create_sheet("가이드")

    guide_content = [
        ("전문가 지식 입력 가이드", header_font, header_fill),
        ("", None, None),
        ("1. 인과관계 시트", Font(bold=True), subheader_fill),
        ("   - 물리적/공학적으로 확실한 인과관계를 입력합니다.", None, None),
        ("   - 관계 유형:", None, None),
        ("     • CAUSES: 직접적 원인-결과 (예: RF Power → Etch Rate)", None, None),
        ("     • INFLUENCES: 간접적 영향 (예: 온도 → 균일도)", None, None),
        ("     • CORRELATES: 상관관계만 있음 (원인 관계 불명확)", None, None),
        ("     • INHIBITS: 억제 관계 (A 증가 시 B 감소)", None, None),
        ("   - 시간 지연: 원인에서 결과까지 걸리는 시간 (초)", None, None),
        ("   - 신뢰도: 관계의 확실성 (1.0 = 100% 확실)", None, None),
        ("", None, None),
        ("2. 알람 원인 시트", Font(bold=True), subheader_fill),
        ("   - 각 알람이 발생하는 가능한 원인들을 나열합니다.", None, None),
        ("   - 동일 알람에 여러 원인이 있으면 여러 행으로 입력합니다.", None, None),
        ("   - 발생 확률은 경험적 빈도 기반으로 입력합니다.", None, None),
        ("   - 우선순위는 점검 순서를 의미합니다 (1이 가장 먼저).", None, None),
        ("", None, None),
        ("3. 선행 지표 시트", Font(bold=True), subheader_fill),
        ("   - 특정 이벤트 발생 전에 나타나는 징후를 입력합니다.", None, None),
        ("   - 패턴 유형: increasing_trend, spike, drift, oscillation 등", None, None),
        ("   - 선행 시간: 이벤트 발생 몇 초 전부터 감지 가능한지", None, None),
        ("   - 이 정보는 예지 정비(Predictive Maintenance)에 활용됩니다.", None, None),
        ("", None, None),
        ("4. 파라미터 그룹 시트", Font(bold=True), subheader_fill),
        ("   - 함께 분석해야 하는 파라미터들을 그룹으로 묶습니다.", None, None),
        ("   - 한 파라미터 이상 시 그룹 내 다른 파라미터도 확인합니다.", None, None),
        ("", None, None),
        ("5. 불가능한 관계 시트", Font(bold=True), subheader_fill),
        ("   - 데이터 분석 시 발견되더라도 물리적으로 불가능한 관계를 정의합니다.", None, None),
        ("   - 이 관계들은 Agent 추론에서 제외됩니다.", None, None),
        ("   - False Positive 제거에 활용됩니다.", None, None),
        ("", None, None),
        ("6. 파라미터 정의 시트", Font(bold=True), subheader_fill),
        ("   - 분석에 사용되는 파라미터의 메타데이터를 정의합니다.", None, None),
        ("   - 정상 범위와 Spec 범위를 구분합니다.", None, None),
        ("   - 데이터 소스(FDC/SPC/MES 등)를 명시합니다.", None, None),
        ("", None, None),
        ("입력 팁", Font(bold=True, color="FF0000"), warning_fill),
        ("   - 초록색 행은 예시입니다. 삭제하거나 수정해도 됩니다.", None, None),
        ("   - 확실한 것부터 입력하세요. 불확실한 것은 신뢰도를 낮게.", None, None),
        ("   - 모든 관계를 입력할 필요 없습니다. 핵심 관계만 충분합니다.", None, None),
        ("   - 데이터 분석 결과와 비교하여 검증됩니다.", None, None),
    ]

    ws7.column_dimensions['A'].width = 100

    for i, (text, font, fill) in enumerate(guide_content, start=1):
        cell = ws7.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)

    # ================================================================
    # 저장
    # ================================================================
    output_dir = os.path.join(os.path.dirname(__file__), '..', 'templates')
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, 'expert_knowledge_template.xlsx')
    wb.save(output_path)

    print(f"템플릿 생성 완료: {output_path}")
    return output_path


if __name__ == '__main__':
    create_template()
